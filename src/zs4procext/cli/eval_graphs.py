import json
from typing import Any, Dict
from zs4procext.parser import KeywordSearching

import click
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from zs4procext.evaluator_graphs import Evaluator_Graphs
import os


@click.command()
@click.argument('reference_file', type=click.Path(exists=True))
@click.argument('test_file', type=click.Path(exists=True))
@click.argument('output_file', type=click.Path())
@click.option('--label-threshold', default=0.7, help='Threshold for matching labels.')
@click.option('--series-threshold', default=0.7, help='Threshold for matching series.')
@click.option('--point-distance-threshold', default=0.1, help='Threshold for matching points based on euclidean distance.')

def main(reference_file: str, test_file: str, output_file: str, label_threshold: float, series_threshold: float, point_distance_threshold: float):
    """
    REFERENCE_FILE: Path to the reference JSON file.
    TEST_FILE: Path to the test JSON file.
    OUTPUT_FILE: Path to the output Excel file.
    """
    eval_graphs(reference_file, test_file, output_file, label_threshold, series_threshold, point_distance_threshold)

def eval_graphs(
    reference_file: str,
    test_file: str,
    output_file: str,
    label_threshold: float,
    series_threshold: float,
    point_distance_threshold: float
) -> None:
    reference_data = Evaluator_Graphs.load_json(reference_file)
    test_data = Evaluator_Graphs.load_json(test_file)
    
    evaluator = Evaluator_Graphs(reference_data=reference_data, threshold=label_threshold, distance_threshold=point_distance_threshold)
    
    # Process plots and get overall results
    results = evaluator.process_plots(test_data)

    # Calculate metrics for labels, series, and points
    label_metrics = evaluator.evaluate(results["Label_TP"], results["Label_FP"], results["Label_FN"])
    series_metrics = evaluator.evaluate(results["Series_TP"], results["Series_FP"], results["Series_FN"])
    point_metrics = evaluator.evaluate(results["Point_TP"], results["Point_FP"], results["Point_FN"])
    skipped_images = evaluator.evaluate (results["Skipped_Images"], results["Total_Images"], results["Skipped_Percent"])

    # Restructure the DataFrame
    data = {
        "Metric": ["FN", "FP", "TP", "New Detections", "Miss Detections", "Incorrect Detections"],
        "Label Metrics": [results["Label_FN"], results["Label_FP"], results["Label_TP"], label_metrics["New detections"], label_metrics["Miss detections"], label_metrics["Incorrect detections"]],
        "Series Metrics": [results["Series_FN"], results["Series_FP"], results["Series_TP"], series_metrics["New detections"], series_metrics["Miss detections"], series_metrics["Incorrect detections"]],
        "Point Metrics": [results["Point_FN"], results["Point_FP"], results["Point_TP"], point_metrics["New detections"], point_metrics["Miss detections"], point_metrics["Incorrect detections"]],
        "Label Threshold": [label_threshold, "", "", "", "", ""],
        "Series Threshold": [series_threshold, "", "", "", "", ""],
        "Point Distance Threshold": [point_distance_threshold, "", "", "", "", ""],
        "Images Skipped":[results ["Skipped_Images"], "", "", "", "", ""],
        "Total Images": [results ["Total_Images"], "", "", "", "", ""],
        "% Images Skipped": [results ["Skipped_Percent"], "", "", "", "", ""]
    }

    df = pd.DataFrame(data)
    
    # Ensure the output file has a proper extension and filename
    if os.path.isdir(output_file):
        output_file = os.path.join(output_file, 'results.xlsx')
    elif not output_file.endswith(('.xlsx', '.xls')):
        output_file += '.xlsx'
    
    df.to_excel(output_file, index=False)

    # Apply color and bold formatting to specific rows in the Excel file
    wb = load_workbook(output_file)
    ws = wb.active
    
    bold_font = Font(bold=True)
    
    # Apply bold formatting to Precision, Recall, and F-Score rows
    for row in [4, 5, 6]:  # Rows to bold (New Detections, Miss Detections, Incorrect Detections)
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row + 1, column=col)  # +1 because Excel is 1-indexed
            cell.font = bold_font
    
    # Apply gradient color to numeric cells
    for row in range(5, ws.max_row + 1):
        for col in range(2, 5):  # Columns with numeric values
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                if cell.value >= 0.80:
                    fill = PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid")  # Pastel red
                elif cell.value >= 0.70:
                    fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Light red
                elif cell.value >= 0.45:
                    fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Light yellow
                elif cell.value >= 0.10:
                    fill = PatternFill(start_color="D4E9D6", end_color="D4E9D6", fill_type="solid")  # Lighter green
                else:
                    fill = PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid")  # Pastel green
                cell.fill = fill
    
    # Save the workbook
    wb.save(output_file)

if __name__ == "__main__":
    main()
